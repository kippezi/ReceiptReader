from openpyxl import Workbook
from openpyxl import load_workbook
import os

def insert_receipts_to_excel(month_recs,wb_save_location):

    # Open the file if it exists, make anew if it doesn't
    if os.path.isfile(wb_save_location):
        wb = load_workbook(wb_save_location)
    else:
        wb = Workbook()

    # Check whether the month_rec already has a designated sheet in the file, if yes, use that, if no, make anew
    for month_rec in month_recs:
        month_rec_exists = False
        for sheetname in wb.sheetnames:
            if month_rec.month.lower() in sheetname.lower():
                month_rec_exists = True
                break
        if month_rec_exists:
            print("Already in sheetnames " + sheetname)
            ws = wb[sheetname]
        else:
            print("Sheet name: " + sheetname)
            print("Not yet in months: " + month_rec.month.lower())
            ws = wb.create_sheet(month_rec.month.capitalize())

        # Write the column explanation
        ws['A1'] = 'Päivämäärä'
        ws['B1'] = 'Tapahtuman Selite'
        ws['C1'] = 'Kuitti ID'
        ws['D1'] = 'Tulo/Meno'
        ws['E1'] = 'Luokka'
        ws['F1'] = 'Hinta'
        ws['F1'] = 'Valuutta'

        # Write the actual data from receipts
        for receipt in month_rec.receipts:
            idx = find_first_empty_row(ws)
            ws['A' + str(idx)] = receipt.receipt_date.strftime("%d.%m.%Y")
            ws['B' + str(idx)] = receipt.explanation
            ws['C' + str(idx)] = receipt.id
            ws['D' + str(idx)] = receipt.receipt_type
            ws['F' + str(idx)] = receipt.amount
            ws['G' + str(idx)] = receipt.currency

            if receipt.tax_cat is not None:
                ws['E' + str(idx)] = receipt.tax_cat

    if os.path.isfile(wb_save_location):
        wb.save(wb_save_location)
    else:
        wb.save(wb_save_location + r"\receipts_read.xlsx")

def find_first_empty_row(ws):
    idx = 2
    while True:
        if ws['A' + str(idx)].value is None or idx > 100:
            return idx
            break
        idx += 1



